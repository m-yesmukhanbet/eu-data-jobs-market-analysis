"""
Build eu_da_market_onepager.xlsx from pbi_jobs.csv and pbi_skills_long.csv.
Loads raw rows into Excel and uses Excel formulas (not Python-computed hardcodes)
for all summary numbers, pivots, and the chart.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


jobs = pd.read_csv('pbi_jobs.csv')
skills = pd.read_csv('pbi_skills_long.csv')

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color='1F4E78')
LABEL_FONT = Font(name=FONT_NAME, size=11, color='595959')
BIG_NUMBER_FONT = Font(name=FONT_NAME, bold=True, size=20, color='1F4E78')
NORMAL_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------- Sheet: RawJobs (hidden helper data) ----------
raw = wb.active
raw.title = 'RawJobs'
job_cols = ['id', 'title', 'company', 'country', 'city', 'created',
            'salary_eur', 'contract_time', 'is_junior', 'is_remote']
raw.append(job_cols)
for cell in raw[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
for _, row in jobs[job_cols].iterrows():
    raw.append(list(row))
n_jobs = len(jobs)
raw.sheet_state = 'hidden'

# ---------- Sheet: RawSkills (hidden helper data) ----------
raw_sk = wb.create_sheet('RawSkills')
raw_sk.append(['job_id', 'skill', 'country'])
for cell in raw_sk[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
for _, row in skills.iterrows():
    raw_sk.append(list(row))
n_skills = len(skills)
raw_sk.sheet_state = 'hidden'

country_col = job_cols.index('country') + 1  # D
salary_col = job_cols.index('salary_eur') + 1  # G
is_junior_col = job_cols.index('is_junior') + 1  # I
JOBS_RANGE_COUNTRY = f'RawJobs!${get_column_letter(country_col)}$2:${get_column_letter(country_col)}${n_jobs+1}'
JOBS_RANGE_SALARY = f'RawJobs!${get_column_letter(salary_col)}$2:${get_column_letter(salary_col)}${n_jobs+1}'
JOBS_RANGE_JUNIOR = f'RawJobs!${get_column_letter(is_junior_col)}$2:${get_column_letter(is_junior_col)}${n_jobs+1}'

SK_RANGE_SKILL = f'RawSkills!$B$2:$B${n_skills+1}'

# ---------- Sheet: Summary ----------
ws = wb.create_sheet('Summary', 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
for col, w in zip('BCDEF', [26, 20, 20, 20, 20]):
    ws.column_dimensions[col].width = w

ws['B2'] = 'EU Data Analyst Job Market — Summary'
ws['B2'].font = TITLE_FONT
ws.merge_cells('B2:F2')

ws['B3'] = 'Adzuna postings, July 2026 snapshot'
ws['B3'].font = LABEL_FONT
ws.merge_cells('B3:F3')

kpi_labels = ['Total Postings', '% with Salary Data', 'Median Salary EUR (GB)',
              'Median Salary EUR (FR)', 'Top Skill (by mentions)']
kpi_cells = ['B', 'C', 'D', 'E', 'F']
row_label, row_value = 5, 6

for col, label in zip(kpi_cells, kpi_labels):
    c = ws[f'{col}{row_label}']
    c.value = label
    c.font = LABEL_FONT
    c.alignment = Alignment(horizontal='center', wrap_text=True)

ws[f'B{row_value}'] = f'=COUNTA({JOBS_RANGE_COUNTRY})'
ws[f'C{row_value}'] = f'=COUNT({JOBS_RANGE_SALARY})/COUNTA({JOBS_RANGE_COUNTRY})'
ws[f'C{row_value}'].number_format = '0.0%'
d_formula = f'=MEDIAN(IF(RawJobs!$D$2:$D${n_jobs+1}="gb",{JOBS_RANGE_SALARY}))'
e_formula = f'=MEDIAN(IF(RawJobs!$D$2:$D${n_jobs+1}="fr",{JOBS_RANGE_SALARY}))'
ws[f'D{row_value}'] = ArrayFormula(f'D{row_value}', d_formula)
ws[f'E{row_value}'] = ArrayFormula(f'E{row_value}', e_formula)
ws[f'D{row_value}'].number_format = '#,##0'
ws[f'E{row_value}'].number_format = '#,##0'
f_formula = f'=INDEX({SK_RANGE_SKILL},MATCH(MAX(COUNTIF({SK_RANGE_SKILL},{SK_RANGE_SKILL})),COUNTIF({SK_RANGE_SKILL},{SK_RANGE_SKILL}),0))'
ws[f'F{row_value}'] = ArrayFormula(f'F{row_value}', f_formula)

for col in kpi_cells:
    c = ws[f'{col}{row_value}']
    c.font = BIG_NUMBER_FONT
    c.alignment = Alignment(horizontal='center')

# Country counts for chart (Excel formulas, not hardcoded)
ws['B9'] = 'Postings by Country'
ws['B9'].font = Font(name=FONT_NAME, bold=True, size=12, color='1F4E78')
ws.merge_cells('B9:F9')

countries = sorted(jobs['country'].unique().tolist())
ws['B10'] = 'Country'
ws['C10'] = 'Postings'
for c in ['B10', 'C10']:
    ws[c].font = HEADER_FONT
    ws[c].fill = HEADER_FILL

for i, ctry in enumerate(countries):
    r = 11 + i
    ws[f'B{r}'] = ctry.upper()
    ws[f'C{r}'] = f'=COUNTIF({JOBS_RANGE_COUNTRY},B{r})'
    ws[f'B{r}'].font = NORMAL_FONT
    ws[f'C{r}'].font = NORMAL_FONT
    ws[f'B{r}'].border = BORDER
    ws[f'C{r}'].border = BORDER

last_country_row = 10 + len(countries)

chart = BarChart()
chart.type = 'col'
chart.title = 'Job Postings by Country'
chart.y_axis.title = 'Postings'
chart.x_axis.title = 'Country'
chart.style = 10
data = Reference(ws, min_col=3, min_row=10, max_row=last_country_row)
cats = Reference(ws, min_col=2, min_row=11, max_row=last_country_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 16
chart.height = 9
ws.add_chart(chart, f'B{last_country_row + 2}')

# ---------- Sheet: Pivot_Country ----------
ws2 = wb.create_sheet('Pivot_Country')
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 20

ws2['A1'] = 'Country'
ws2['B1'] = 'Postings'
ws2['C1'] = 'Median Salary EUR'
for c in ['A1', 'B1', 'C1']:
    ws2[c].font = HEADER_FONT
    ws2[c].fill = HEADER_FILL

for i, ctry in enumerate(countries):
    r = 2 + i
    ws2[f'A{r}'] = ctry.upper()
    ws2[f'B{r}'] = f'=COUNTIF({JOBS_RANGE_COUNTRY},A{r})'
    c_formula = f'=IFERROR(MEDIAN(IF(RawJobs!$D$2:$D${n_jobs+1}=A{r},{JOBS_RANGE_SALARY})),"n/a")'
    ws2[f'C{r}'] = ArrayFormula(f'C{r}', c_formula)
    ws2[f'C{r}'].number_format = '#,##0'
    for col in 'ABC':
        ws2[f'{col}{r}'].font = NORMAL_FONT
        ws2[f'{col}{r}'].border = BORDER

# ---------- Sheet: Pivot_Skills ----------
ws3 = wb.create_sheet('Pivot_Skills')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 14

ws3['A1'] = 'Skill'
ws3['B1'] = 'Mentions'
for c in ['A1', 'B1']:
    ws3[c].font = HEADER_FONT
    ws3[c].fill = HEADER_FILL

top_skills = skills['skill'].value_counts().head(15).index.tolist()
for i, sk in enumerate(top_skills):
    r = 2 + i
    ws3[f'A{r}'] = sk
    ws3[f'B{r}'] = f'=COUNTIF({SK_RANGE_SKILL},A{r})'
    ws3[f'A{r}'].font = NORMAL_FONT
    ws3[f'B{r}'].font = NORMAL_FONT
    ws3[f'A{r}'].border = BORDER
    ws3[f'B{r}'].border = BORDER

wb.save('eu_da_market_onepager.xlsx')
print('saved')
